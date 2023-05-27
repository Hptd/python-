import math
import time
import openpyxl
from multiprocessing import Process, Queue


class Test(object):
    def __init__(self):
        self.row_list = []
        self.file_name = "./源文件/SPOUSE_filter_id.xlsx"

    def test_func(self):
        start_time = time.time()

        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.active

        print(ws.max_row)  # type: ignore

        total = ws.max_row  # type: ignore
        psize = 1000
        count_page = math.ceil(total / psize)

        q = Queue()
        t_obj = []
        for i in range(count_page):
            print(i)
            start = i * psize if i else 2
            end = min((i+1) * psize, ws.max_row + 1) # type: ignore
            t = Process(target=self.get_delete_list_for_func, args=(ws, start, end, q))
            t_obj.append(t)
            t.start()
        
        for t in t_obj:
            t.join()

        while not q.empty():
            self.row_list.extend(q.get())

        # self.get_delete_list_for_func(ws, 2, ws.max_row + 1) # type: ignore

        print(self.row_list)

        end_time = time.time()
        print("耗时：", end_time - start_time)

    def get_delete_list_for_func(self, ws, row_min, row_max, q):
        delete_list = []
        for row in range(row_min, row_max):
            # print("\r", row, end="")
            # 统计一行有多少个int类型值的单元格
            count = 0
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value == 0 or ws.cell(row=row, column=col).value == 1:
                    count += 1
            if count <= 14:
                delete_list.append(row)
        print(delete_list)
        q.put(delete_list)
if __name__ == '__main__':
    Test().test_func()
